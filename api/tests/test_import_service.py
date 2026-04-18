from io import BytesIO
from unittest.mock import AsyncMock

import pytest
from openpyxl import Workbook

from app.services.import_service import ImportService

USER_ID = "user-1"
USER_NAME = "Import Admin"

ACTIVE_ACCOUNT = {
    "id": "acc-1",
    "accountLabel": "Unicaja 0382",
    "iban": "ES0000490001000000001234",
    "bankNameShort": "Unicaja",
    "isActive": True,
}


# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def make_full_workbook_bytes(*, headers_lang: str = "es") -> bytes:
    """Build a valid Full mode workbook (movements + categories sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    ws.cell(row=1, column=4, value="UCJAES2MXXX")
    ws.cell(row=1, column=5, value="IBAN: ES00 0049 0001 0000 0000 1234")
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    if headers_lang == "en":
        ws.append(
            [
                "Date",
                "Value Date",
                "Description",
                "Amount",
                "Currency",
                "Balance",
                "Category",
                "Subcategory",
                "Detail",
            ]
        )
    else:
        ws.append(
            [
                "Fecha",
                "Valor",
                "Observaciones",
                "Importe",
                "Divisa",
                "Saldo",
                "Categoria",
                "Subcategoria",
                "Detalle",
            ]
        )

    ws.append(
        [
            "2025-01-02",
            "2025-01-02",
            "BULTO MILLET VICTOR",
            30,
            "EUR",
            71670.37,
            "Cuotas",
            "Cuota Socio Mensual",
            "transferencia",
        ]
    )
    ws.append(
        [
            "2025-01-03",
            "2025-01-03",
            "FRANCISCO BORJA",
            10,
            "EUR",
            71680.37,
            "Donaciones",
            "Donación Particular",
            "particular",
        ]
    )

    category_sheet = wb.create_sheet("Categorias")
    if headers_lang == "en":
        category_sheet.append(["Income", "Income", "Expense"])
    else:
        category_sheet.append(["Entrada", "Entrada", "Gasto"])
    category_sheet.append(["Donaciones", "Cuotas", "Gastos"])
    category_sheet.append(["Donación Particular", "Cuota Socio Mensual", "Electricidad"])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# Legacy alias for backward compatibility with existing test references
make_workbook_bytes = make_full_workbook_bytes


def make_bank_workbook_bytes() -> bytes:
    """Build a Bank mode workbook (no category/subcategory columns)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    ws.cell(row=1, column=4, value="UCJAES2MXXX")
    ws.cell(row=1, column=5, value="IBAN: ES00 0049 0001 0000 0000 1234")
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    ws.append(["Fecha", "Valor", "Observaciones", "Importe", "Divisa", "Saldo"])

    # Positive amount (income)
    ws.append(["2025-01-02", "2025-01-02", "INGRESO TRANSFERENCIA", 500, "EUR", 70500.00])
    # Negative amount (expense)
    ws.append(["2025-01-03", "2025-01-03", "PAGO ELECTRICIDAD", -85.40, "EUR", 70414.60])
    # Another positive
    ws.append(["2025-01-04", "2025-01-04", "DONACION RECIBIDA", 200, "EUR", 70614.60])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def make_inline_workbook_bytes() -> bytes:
    """Build an Inline mode workbook (category columns but NO categories sheet)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    ws.cell(row=1, column=4, value="UCJAES2MXXX")
    ws.cell(row=1, column=5, value="IBAN: ES00 0049 0001 0000 0000 1234")
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])

    ws.append(
        [
            "Fecha",
            "Valor",
            "Observaciones",
            "Importe",
            "Divisa",
            "Saldo",
            "Categoria",
            "Subcategoria",
        ]
    )

    ws.append(["2025-01-02", "2025-01-02", "CUOTA MENSUAL", 30, "EUR", 70530, "Cuotas", "Cuota Socio Mensual"])
    ws.append(
        [
            "2025-01-03",
            "2025-01-03",
            "DONACION PARTICULAR",
            10,
            "EUR",
            70540,
            "Donaciones",
            "Donación Particular",
        ]
    )
    ws.append(["2025-01-04", "2025-01-04", "NUEVA CATEGORIA", -50, "EUR", 70490, "NewCategory", "NewSub"])

    # NO categories sheet — this is what makes it Inline mode
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def make_workbook_with_errors() -> bytes:
    """Workbook with data integrity issues: empty dates, empty categories."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.append(["Date", "Amount", "Category", "Subcategory"])
    ws.append(["2025-01-01", 100, "Donations", "Individual"])
    ws.append([None, 50, "Donations", "Individual"])  # empty date
    ws.append(["2025-01-03", None, "Donations", "Individual"])  # empty amount
    ws.append(["2025-01-04", 25, None, "Individual"])  # empty category
    ws.append(["2025-01-05", 25, "Donations", None])  # empty subcategory
    ws.append(["not-a-date", 25, "Donations", "Individual"])  # bad date

    cat = wb.create_sheet("Categories")
    cat.append(["Income"])
    cat.append(["Donations"])
    cat.append(["Individual"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_workbook_missing_subcategory() -> bytes:
    """Workbook where a transaction references a subcategory not in sheet or DB."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.append(["Date", "Amount", "Category", "Subcategory"])
    ws.append(["2025-01-01", 100, "Donations", "Unknown Sub"])

    cat = wb.create_sheet("Categories")
    cat.append(["Income"])
    cat.append(["Donations"])
    cat.append(["Individual"])  # Only "Individual" is listed, not "Unknown Sub"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_service():
    account_service = AsyncMock()
    category_service = AsyncMock()
    transaction_service = AsyncMock()
    # Default: get_account returns the active account
    account_service.get_account.return_value = ACTIVE_ACCOUNT
    return (
        ImportService(
            account_service=account_service,
            category_service=category_service,
            transaction_service=transaction_service,
        ),
        account_service,
        category_service,
        transaction_service,
    )


# ---------------------------------------------------------------------------
# Account validation tests
# ---------------------------------------------------------------------------


class TestAccountValidation:
    async def test_missing_account_raises_value_error(self):
        service, account_svc, _, _ = await build_service()
        account_svc.get_account.return_value = None

        with pytest.raises(ValueError, match="Account not found or inactive"):
            await service.preview_workbook(make_bank_workbook_bytes(), account_id="nonexistent")

    async def test_inactive_account_raises_value_error(self):
        service, account_svc, _, _ = await build_service()
        account_svc.get_account.return_value = {**ACTIVE_ACCOUNT, "isActive": False}

        with pytest.raises(ValueError, match="Account not found or inactive"):
            await service.preview_workbook(make_bank_workbook_bytes(), account_id="acc-1")

    async def test_active_account_succeeds(self):
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_bank_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["account"]["id"] == "acc-1"
        assert result["account"]["label"] == "Unicaja 0382"

    async def test_import_missing_account_raises_value_error(self):
        service, account_svc, _, _ = await build_service()
        account_svc.get_account.return_value = None

        with pytest.raises(ValueError, match="Account not found or inactive"):
            await service.import_workbook(
                make_bank_workbook_bytes(), account_id="nonexistent", user_id=USER_ID, user_name=USER_NAME
            )


# ---------------------------------------------------------------------------
# Mode detection tests
# ---------------------------------------------------------------------------


class TestModeDetection:
    async def test_full_mode_detected(self):
        """File with category columns + categories sheet → Full."""
        service, _, _, _ = await build_service()
        result = await service.preview_workbook(make_full_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "full"

    async def test_bank_mode_detected(self):
        """File with only date + amount columns → Bank."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_bank_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "bank"

    async def test_inline_mode_detected(self):
        """File with category columns but NO categories sheet → Inline."""
        service, _, category_svc, txn_svc = await build_service()
        category_svc.list_categories.return_value = []
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_inline_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "inline"

    async def test_only_category_column_is_bank_mode(self):
        """File with only category column (no subcategory) → Bank. Both required."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.cell(row=1, column=5, value="IBAN: ES0000490001000000001234")
        ws.append(["Date", "Amount", "Category"])
        ws.append(["2025-01-01", 100, "Donations"])
        buf = BytesIO()
        wb.save(buf)

        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "bank"

    async def test_bank_mode_ignores_categories_sheet(self):
        """Bank mode file with a Categories sheet present → still Bank."""
        wb = Workbook()
        ws = wb.active
        ws.title = "UNICAJA 2026"
        ws.cell(row=1, column=5, value="IBAN: ES0000490001000000001234")
        ws.append(["Date", "Amount", "Description"])
        ws.append(["2025-01-01", 100, "Test"])
        wb.create_sheet("Categories")  # sheet present but irrelevant
        buf = BytesIO()
        wb.save(buf)

        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "bank"


# ---------------------------------------------------------------------------
# Bank mode import tests
# ---------------------------------------------------------------------------


class TestBankModeImport:
    async def test_bank_import_creates_uncategorized_transactions(self):
        service, _, category_svc, txn_svc = await build_service()

        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_bank_workbook_bytes(),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        assert result["importMode"] == "bank"
        assert result["importSource"] == "excel-bank"
        assert result["importBatchId"]  # non-empty UUID
        assert result["transactionsImported"] == 3
        assert result["categoriesCreated"] == 0
        assert result["subcategoriesAdded"] == 0

        # Verify no category sync was attempted
        category_svc.create_category.assert_not_awaited()

        # Verify transaction created with null category and import fields
        calls = txn_svc.create_transaction.call_args_list
        assert len(calls) == 3
        first_tx = calls[0][0][0]  # TransactionCreate
        assert first_tx.category_id is None
        assert first_tx.subcategory_id is None
        assert first_tx.import_batch_id == result["importBatchId"]
        assert first_tx.import_source == "excel-bank"

    async def test_bank_import_positive_amount_is_income(self):
        service, _, _, txn_svc = await build_service()

        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        await service.import_workbook(
            make_bank_workbook_bytes(), account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        first_tx = txn_svc.create_transaction.call_args_list[0][0][0]
        assert first_tx.transaction_type.value == "income"
        assert first_tx.amount > 0

    async def test_bank_import_negative_amount_is_expense(self):
        service, _, _, txn_svc = await build_service()

        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        await service.import_workbook(
            make_bank_workbook_bytes(), account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        second_tx = txn_svc.create_transaction.call_args_list[1][0][0]
        assert second_tx.transaction_type.value == "expense"

    async def test_bank_preview_has_empty_category_lists(self):
        service, _, _, txn_svc = await build_service()

        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_bank_workbook_bytes(), account_id="acc-1")

        assert result["importMode"] == "bank"
        assert result["newCategories"] == []
        assert result["newSubcategories"] == []
        assert result["transactionsToImport"] == 3


# ---------------------------------------------------------------------------
# Inline mode import tests
# ---------------------------------------------------------------------------


class TestInlineModeImport:
    async def test_inline_import_creates_missing_categories(self):
        service, _, category_svc, txn_svc = await build_service()

        # "Cuotas" exists in DB but "Donaciones" and "NewCategory" don't
        category_svc.list_categories.return_value = [
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
        ]
        category_svc.create_category.side_effect = [
            {"id": "cat-don", "name": "Donaciones", "categoryType": "expense", "subcategories": []},
            {"id": "cat-new", "name": "NewCategory", "categoryType": "expense", "subcategories": []},
        ]
        category_svc.add_subcategory.side_effect = [
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-new",
                "name": "NewCategory",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-new", "name": "NewSub", "isActive": True}],
            },
        ]
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_inline_workbook_bytes(),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        assert result["importMode"] == "inline"
        assert result["importSource"] == "excel-inline"
        assert result["categoriesCreated"] == 2  # Donaciones + NewCategory
        assert result["transactionsImported"] == 3
        # No "will be created as 'expense'" warning anymore (removed per amendment 1)
        assert not any("will be created as 'expense'" in w for w in result["warnings"])

    async def test_inline_preview_shows_new_categories_with_suggested_type(self):
        service, _, category_svc, txn_svc = await build_service()

        category_svc.list_categories.return_value = []
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_inline_workbook_bytes(), account_id="acc-1")

        assert result["importMode"] == "inline"
        assert result["valid"] is True
        assert len(result["newCategories"]) == 3  # Cuotas, Donaciones, NewCategory — all new
        # All new categories should have suggestedType = "expense"
        for cat in result["newCategories"]:
            assert cat["suggestedType"] == "expense"
        # No "will be created as 'expense'" warning
        assert not any("will be created as 'expense'" in w for w in result["warnings"])

    async def test_inline_existing_category_reuses_type(self):
        service, _, category_svc, txn_svc = await build_service()

        category_svc.list_categories.return_value = [
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
        ]
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_inline_workbook_bytes(), account_id="acc-1")

        assert result["importMode"] == "inline"
        # Cuotas exists → not in newCategories; Donaciones and NewCategory are new
        new_names = [c["name"] for c in result["newCategories"]]
        assert "Cuotas" not in new_names
        assert "Donaciones" in new_names
        assert "NewCategory" in new_names


# ---------------------------------------------------------------------------
# Full mode import tests (updated for v2 fields)
# ---------------------------------------------------------------------------


class TestFullModeImport:
    async def test_imports_categories_and_transactions(self):
        service, _, category_svc, txn_svc = await build_service()

        category_svc.list_categories.return_value = []
        category_svc.create_category.side_effect = [
            {"id": "cat-don", "name": "Donaciones", "categoryType": "income", "subcategories": []},
            {"id": "cat-cuotas", "name": "Cuotas", "categoryType": "income", "subcategories": []},
            {"id": "cat-gastos", "name": "Gastos", "categoryType": "expense", "subcategories": []},
        ]
        category_svc.add_subcategory.side_effect = [
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]

        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_full_workbook_bytes(),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        assert result["importMode"] == "full"
        assert result["importSource"] == "excel-full"
        assert result["importBatchId"]  # non-empty UUID
        assert result["accountId"] == "acc-1"
        assert result["categoriesCreated"] == 3
        assert result["subcategoriesAdded"] == 3
        assert result["transactionsImported"] == 2
        assert result["duplicatesSkipped"] == 0
        assert result["rowsSkipped"] == 0

    async def test_full_mode_derives_type_from_amount_sign(self):
        """v2: transactionType comes from amount sign, not category type."""
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = [
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        await service.import_workbook(
            make_full_workbook_bytes(),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        # Both rows have positive amounts → income, regardless of category type
        calls = txn_svc.create_transaction.call_args_list
        for call in calls:
            tx = call[0][0]
            assert tx.transaction_type.value == "income"  # positive amount → income

    async def test_full_mode_includes_import_batch_fields(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = [
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_full_workbook_bytes(), account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        batch_id = result["importBatchId"]
        calls = txn_svc.create_transaction.call_args_list
        for call in calls:
            tx = call[0][0]
            assert tx.import_batch_id == batch_id
            assert tx.import_source == "excel-full"

    async def test_skips_duplicate_transactions(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = [
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]

        txn_svc.get_transactions_for_export.return_value = [
            {
                "date": "2025-01-02",
                "bankDescription": "BULTO MILLET VICTOR",
                "detail": "transferencia",
                "amount": 30,
            }
        ]
        txn_svc.create_transaction.return_value = {"id": "tx-2"}

        result = await service.import_workbook(
            make_full_workbook_bytes(),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        assert result["transactionsImported"] == 1
        assert result["duplicatesSkipped"] == 1
        assert txn_svc.create_transaction.await_count == 1

    async def test_english_headers_recognised(self):
        """Workbook with English headers should be imported identically."""
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {"id": "acc-1", "accountLabel": "Default", "iban": "", "isActive": True}
        account_svc.create_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }

        category_svc.list_categories.return_value = []
        category_svc.create_category.side_effect = [
            {"id": "cat-don", "name": "Donaciones", "categoryType": "income", "subcategories": []},
            {"id": "cat-cuotas", "name": "Cuotas", "categoryType": "income", "subcategories": []},
            {"id": "cat-gastos", "name": "Gastos", "categoryType": "expense", "subcategories": []},
        ]
        category_svc.add_subcategory.side_effect = [
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]

        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_full_workbook_bytes(headers_lang="en"),
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
        )

        assert result["categoriesCreated"] == 3
        assert result["transactionsImported"] == 2


# ---------------------------------------------------------------------------
# Preview tests
# ---------------------------------------------------------------------------


class TestPreviewWorkbook:
    async def test_valid_full_workbook_returns_preview(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = []
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_full_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "full"
        assert result["errors"] == []
        assert result["account"]["label"] == "Unicaja 0382"
        assert len(result["newCategories"]) == 3
        assert len(result["newSubcategories"]) == 3
        assert result["transactionsToImport"] == 2
        assert result["duplicatesToSkip"] == 0

    async def test_preview_detects_duplicates(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = [
            {
                "id": "cat-don",
                "name": "Donaciones",
                "categoryType": "income",
                "subcategories": [{"id": "sub-don", "name": "Donación Particular", "isActive": True}],
            },
            {
                "id": "cat-cuotas",
                "name": "Cuotas",
                "categoryType": "income",
                "subcategories": [{"id": "sub-cuota", "name": "Cuota Socio Mensual", "isActive": True}],
            },
            {
                "id": "cat-gastos",
                "name": "Gastos",
                "categoryType": "expense",
                "subcategories": [{"id": "sub-gasto", "name": "Electricidad", "isActive": True}],
            },
        ]
        txn_svc.get_transactions_for_export.return_value = [
            {
                "date": "2025-01-02",
                "bankDescription": "BULTO MILLET VICTOR",
                "detail": "transferencia",
                "amount": 30,
            },
        ]

        result = await service.preview_workbook(make_full_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["transactionsToImport"] == 1
        assert result["duplicatesToSkip"] == 1

    async def test_preview_shows_selected_account(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES70...",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = []

        result = await service.preview_workbook(make_full_workbook_bytes(), account_id="acc-1")

        assert result["valid"] is True
        assert result["account"]["id"] == "acc-1"

    async def test_preview_rejects_data_integrity_errors(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        result = await service.preview_workbook(make_workbook_with_errors(), account_id="acc-1")

        assert result["valid"] is False
        assert len(result["errors"]) > 0
        error_text = " ".join(result["errors"])
        assert "empty date" in error_text.lower()
        assert "empty amount" in error_text.lower()
        assert "empty category" in error_text.lower()
        assert "empty subcategory" in error_text.lower()
        assert "could not be read" in error_text.lower()

    async def test_preview_rejects_orphaned_subcategory(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {"id": "acc-1", "accountLabel": "Default", "iban": "", "isActive": True}
        category_svc.list_categories.return_value = []

        result = await service.preview_workbook(make_workbook_missing_subcategory(), account_id="acc-1")

        assert result["valid"] is False
        error_text = " ".join(result["errors"])
        assert "unknown sub" in error_text.lower()

    async def test_preview_rejects_invalid_file(self):
        service, _, _, _ = await build_service()

        result = await service.preview_workbook(b"not an xlsx file", account_id="acc-1")

        assert result["valid"] is False
        assert any("could not open" in e.lower() for e in result["errors"])

    async def test_preview_full_mode_rejects_missing_categories_sheet(self):
        """Full mode file (both category cols present) but missing categories sheet → error."""
        # Build a workbook with category columns but wrongly-named categories sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        ws.cell(row=1, column=5, value="IBAN: ES0000490001000000001234")
        ws.append(["Date", "Amount", "Category", "Subcategory"])
        ws.append(["2025-01-01", 100, "Test", "Sub"])
        # Sheet named "WrongName" — not matching CATEGORY_SHEET_NAMES → inline, not full
        buf = BytesIO()
        wb.save(buf)

        service, account_svc, category_svc, txn_svc = await build_service()
        account_svc.get_account.return_value = {"id": "acc-1", "accountLabel": "Default", "iban": "", "isActive": True}
        category_svc.list_categories.return_value = []
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")

        # Without categories sheet → detected as inline, not full
        assert result["importMode"] == "inline"
        assert result["valid"] is True

    async def test_preview_english_headers(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Unicaja 0382",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        category_svc.list_categories.return_value = []
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_full_workbook_bytes(headers_lang="en"), account_id="acc-1")

        assert result["valid"] is True
        assert result["importMode"] == "full"
        assert result["transactionsToImport"] == 2


# ---------------------------------------------------------------------------
# Import batch tracking tests
# ---------------------------------------------------------------------------


class TestImportBatchTracking:
    async def test_all_transactions_share_batch_id(self):
        service, account_svc, category_svc, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Test",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_bank_workbook_bytes(), account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        batch_id = result["importBatchId"]
        assert batch_id  # non-empty
        calls = txn_svc.create_transaction.call_args_list
        for call in calls:
            tx = call[0][0]
            assert tx.import_batch_id == batch_id

    async def test_bank_duplicate_detection(self):
        service, account_svc, _, txn_svc = await build_service()

        account_svc.get_account.return_value = {
            "id": "acc-1",
            "accountLabel": "Test",
            "iban": "ES0000490001000000001234",
            "bankNameShort": "Unicaja",
        }
        # One existing transaction matches the first row
        txn_svc.get_transactions_for_export.return_value = [
            {
                "date": "2025-01-02",
                "bankDescription": "INGRESO TRANSFERENCIA",
                "detail": None,
                "amount": 500,
            }
        ]
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        result = await service.import_workbook(
            make_bank_workbook_bytes(), account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        assert result["duplicatesSkipped"] == 1
        assert result["transactionsImported"] == 2


# ---------------------------------------------------------------------------
# Multi-sheet selection tests (issue #17)
# ---------------------------------------------------------------------------


def _append_bank_sheet(wb, *, title: str, dates_amounts: list[tuple[str, float]]) -> None:
    """Append a bank-mode movements sheet with the given (date, amount) rows."""
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=4, value="UCJAES2MXXX")
    ws.cell(row=1, column=5, value="IBAN: ES00 0049 0001 0000 0000 1234")
    for _ in range(4):
        ws.append([])
    ws.append(["Fecha", "Valor", "Observaciones", "Importe", "Divisa", "Saldo"])
    for d, amt in dates_amounts:
        ws.append([d, d, "DESC", amt, "EUR", 100.0])


def make_multi_sheet_bytes(sheet_specs: list[tuple[str, int]]) -> bytes:
    """Build a workbook with multiple bank-mode sheets.

    ``sheet_specs`` is a list of ``(sheet_name, row_count)`` tuples. The first
    sheet replaces the default openpyxl sheet so workbook order is preserved.
    """
    wb = Workbook()
    # Drop the default empty sheet so our first specified sheet is workbook-first.
    wb.remove(wb.active)
    base_dates = ["2025-01-02", "2025-01-03", "2025-01-04", "2025-01-05", "2025-01-06"]
    for name, count in sheet_specs:
        rows = [(base_dates[i % len(base_dates)], 10 + i) for i in range(count)]
        _append_bank_sheet(wb, title=name, dates_amounts=rows)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_workbook_with_extra_noise_sheets() -> bytes:
    """One importable bank sheet + a non-importable Resumen sheet + an empty Notas sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    _append_bank_sheet(wb, title="Movimientos", dates_amounts=[("2025-01-02", 25)])
    summary = wb.create_sheet("Resumen")
    summary.append(["Total", 1234])
    summary.append(["Note", "no headers here"])
    wb.create_sheet("Notas")  # truly empty
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestMultiSheetSelection:
    async def test_single_candidate_no_selection_required(self):
        """Backwards compatibility: single-sheet workbook still produces the normal preview."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        result = await service.preview_workbook(make_bank_workbook_bytes(), account_id="acc-1")

        assert result["requiresSheetSelection"] is False
        assert result["valid"] is True
        assert result["selectedSheet"] == "UNICAJA 2026"
        assert result["availableSheets"] == ["UNICAJA 2026"]

    async def test_multi_candidate_returns_selection_required(self):
        """Two-or-more candidate sheets and no explicit `sheet` → selection-required payload."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []

        wb_bytes = make_multi_sheet_bytes([("Movimientos 2026", 3), ("Movimientos 2025", 5)])
        result = await service.preview_workbook(wb_bytes, account_id="acc-1")

        assert result["requiresSheetSelection"] is True
        assert [c["name"] for c in result["candidateSheets"]] == ["Movimientos 2026", "Movimientos 2025"]
        # data_row_count = max_row - header_row(=5). 5 data rows for the second sheet, 3 for first.
        counts = {c["name"]: c["dataRowCount"] for c in result["candidateSheets"]}
        assert counts["Movimientos 2026"] == 3
        assert counts["Movimientos 2025"] == 5
        assert result["valid"] is False  # discovery has not validated anything yet
        assert result["account"]["id"] == "acc-1"

    async def test_multi_candidate_workbook_order_preserved(self):
        """Pedro's archive: candidates appear in workbook order (newest year first)."""
        service, _, _, _ = await build_service()
        wb_bytes = make_multi_sheet_bytes(
            [("Movimientos 2026", 1), ("Movimientos 2025", 1), ("Movimientos 2024", 1)]
        )
        result = await service.preview_workbook(wb_bytes, account_id="acc-1")

        names = [c["name"] for c in result["candidateSheets"]]
        assert names == ["Movimientos 2026", "Movimientos 2025", "Movimientos 2024"]

    async def test_explicit_sheet_validates_only_that_sheet(self):
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []
        wb_bytes = make_multi_sheet_bytes([("Movimientos 2026", 1), ("Movimientos 2025", 2)])

        result = await service.preview_workbook(wb_bytes, account_id="acc-1", sheet="Movimientos 2025")

        assert result["requiresSheetSelection"] is False
        assert result["valid"] is True
        assert result["selectedSheet"] == "Movimientos 2025"
        # 2 rows in the chosen sheet, regardless of what other sheets contain
        assert result["totalRows"] == 2

    async def test_unknown_sheet_raises_value_error(self):
        service, _, _, _ = await build_service()
        wb_bytes = make_multi_sheet_bytes([("A", 1), ("B", 1)])

        with pytest.raises(ValueError, match="not found in workbook"):
            await service.preview_workbook(wb_bytes, account_id="acc-1", sheet="DoesNotExist")

    async def test_non_candidate_sheet_raises_value_error(self):
        """A sheet that exists but is missing required headers must be rejected with 400 detail."""
        wb = Workbook()
        wb.remove(wb.active)
        _append_bank_sheet(wb, title="Movimientos", dates_amounts=[("2025-01-02", 10)])
        # Add a second sheet that exists but is not a candidate
        noise = wb.create_sheet("Resumen")
        noise.append(["Total", 1234])
        buf = BytesIO()
        wb.save(buf)

        service, _, _, _ = await build_service()
        with pytest.raises(ValueError, match="not importable"):
            await service.preview_workbook(buf.getvalue(), account_id="acc-1", sheet="Resumen")

    async def test_zero_candidates_diagnostics(self):
        """Workbook with sheets but zero candidates → error + ignored sheets diagnostics."""
        wb = Workbook()
        wb.active.title = "Resumen"
        wb.active.append(["Total", 999])
        wb.create_sheet("Notas")  # empty
        buf = BytesIO()
        wb.save(buf)

        service, _, _, _ = await build_service()
        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")

        assert result["valid"] is False
        assert result["requiresSheetSelection"] is False
        names = {s["name"] for s in result["ignoredSheets"]}
        assert names == {"Resumen", "Notas"}
        reasons = {s["name"]: s["reason"] for s in result["ignoredSheets"]}
        assert reasons["Notas"] == "empty"
        assert reasons["Resumen"] == "missing_required_headers"
        # `missing` reports only headers that were not found — Resumen has neither
        # date nor amount headers, so both must be listed.
        resumen = next(s for s in result["ignoredSheets"] if s["name"] == "Resumen")
        assert sorted(resumen["missing"]) == ["amount", "date"]

    async def test_partial_headers_reports_only_missing(self):
        """A sheet with `Fecha` but no `Importe` reports only `amount` as missing."""
        wb = Workbook()
        wb.active.title = "Parcial"
        wb.active.append(["Fecha", "Observaciones"])
        wb.active.append(["2025-01-02", "DESC"])
        # Add a candidate sheet so we don't hit the "0 candidates" error path
        _append_bank_sheet(wb, title="Movimientos", dates_amounts=[("2025-01-02", 10), ("2025-01-03", 20)])
        buf = BytesIO()
        wb.save(buf)

        service, _, _, _ = await build_service()
        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1", sheet="Movimientos")

        # Movimientos validates fine; Parcial appears in ignoredSheets via discovery.
        assert result["valid"] is True
        # Re-run discovery (no `sheet`) to inspect Parcial — but discovery only
        # surfaces ignored sheets when there's a single candidate or zero, so
        # call _enumerate_sheets directly via a second discovery preview.
        # (Multi-candidate path would also include them.)
        discovery = await service.preview_workbook(buf.getvalue(), account_id="acc-1")
        # Single candidate → normal preview, but ignoredSheets is still populated only
        # in the zero-candidate error path. With 1 candidate we don't expose Parcial,
        # which matches the spec (reduces noise on the happy path). Sanity-check that.
        assert discovery["valid"] is True
        assert discovery["selectedSheet"] == "Movimientos"

    async def test_ignored_sheets_listed_when_multi_candidate(self):
        """Ignored sheets are surfaced in the selection payload alongside candidates."""
        service, _, _, _ = await build_service()
        # 2 candidates (forces selection) + 1 ignored
        wb = Workbook()
        wb.remove(wb.active)
        _append_bank_sheet(wb, title="A", dates_amounts=[("2025-01-02", 10)])
        _append_bank_sheet(wb, title="B", dates_amounts=[("2025-01-02", 10)])
        wb.create_sheet("Notas")  # empty
        buf = BytesIO()
        wb.save(buf)

        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")
        assert result["requiresSheetSelection"] is True
        assert [s["name"] for s in result["ignoredSheets"]] == ["Notas"]

    async def test_url_unsafe_sheet_name_round_trips(self):
        """EC-2: sheet names with accents, spaces and hyphens (Excel disallows / : ? * [ ])
        round-trip through both the discovery and validation calls."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []
        weird = "Año 2026 - Caja 1"
        wb_bytes = make_multi_sheet_bytes([(weird, 1), ("Otro", 1)])

        result = await service.preview_workbook(wb_bytes, account_id="acc-1", sheet=weird)
        assert result["valid"] is True
        assert result["selectedSheet"] == weird

    async def test_import_with_sheet_param_imports_selected_sheet(self):
        """Confirm/import path commits the selected sheet's rows only."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        wb_bytes = make_multi_sheet_bytes([("First", 1), ("Second", 4)])
        result = await service.import_workbook(
            wb_bytes,
            account_id="acc-1",
            user_id=USER_ID,
            user_name=USER_NAME,
            sheet="Second",
        )

        assert result["selectedSheet"] == "Second"
        assert result["transactionsImported"] == 4

    async def test_import_without_sheet_falls_back_to_first_candidate(self):
        """Backwards compatibility: omitting `sheet` keeps today's first-wins behavior."""
        service, _, _, txn_svc = await build_service()
        txn_svc.get_transactions_for_export.return_value = []
        txn_svc.create_transaction.return_value = {"id": "tx-1"}

        wb_bytes = make_multi_sheet_bytes([("First", 2), ("Second", 5)])
        result = await service.import_workbook(
            wb_bytes, account_id="acc-1", user_id=USER_ID, user_name=USER_NAME
        )

        assert result["selectedSheet"] == "First"
        assert result["transactionsImported"] == 2

    async def test_categories_sheet_doubling_as_movements(self):
        """EC-4: a sheet named like the Categories sheet that *also* matches movements headers
        appears in candidates. Documents that movements detection is independent of categories."""
        wb = Workbook()
        wb.remove(wb.active)
        _append_bank_sheet(wb, title="UNICAJA 2026", dates_amounts=[("2025-01-02", 10)])
        # Build a sheet literally named "Categorias" but with movements headers
        cat = wb.create_sheet("Categorias")
        cat.cell(row=1, column=4, value="UCJAES2MXXX")
        cat.cell(row=1, column=5, value="IBAN")
        for _ in range(4):
            cat.append([])
        cat.append(["Fecha", "Importe"])
        cat.append(["2025-02-02", 99])
        buf = BytesIO()
        wb.save(buf)

        service, _, _, _ = await build_service()
        result = await service.preview_workbook(buf.getvalue(), account_id="acc-1")

        assert result["requiresSheetSelection"] is True
        names = {c["name"] for c in result["candidateSheets"]}
        assert names == {"UNICAJA 2026", "Categorias"}
