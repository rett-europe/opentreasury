"""
Generate test Excel workbooks for the Import feature.

Run:
    python api/tests/fixtures/generate_import_fixtures.py

Output:
    api/tests/fixtures/*.xlsx — one file per scenario
"""

from pathlib import Path

from openpyxl import Workbook

OUTPUT_DIR = Path(__file__).parent


def _save(wb: Workbook, name: str) -> None:
    path = OUTPUT_DIR / name
    wb.save(path)
    print(f"  ✔ {name}")


def _add_metadata(ws, *, iban: str = "IBAN: ES00 0049 0001 0000 0000 1234", swift: str = "UCJAES2MXXX"):
    """Add IBAN/SWIFT metadata to row 1 (Unicaja convention)."""
    ws.cell(row=1, column=4, value=swift)
    ws.cell(row=1, column=5, value=iban)


def _add_categories_sheet_es(wb: Workbook) -> None:
    cs = wb.create_sheet("Categorias")
    cs.append(["Entrada", "Entrada", "Gasto"])
    cs.append(["Donaciones", "Cuotas", "Gastos Generales"])
    cs.append(["Donación Particular", "Cuota Socio Mensual", "Electricidad"])
    cs.append(["Donación Empresa", "Cuota Socio Anual", "Material Oficina"])


def _add_categories_sheet_en(wb: Workbook) -> None:
    cs = wb.create_sheet("Categories")
    cs.append(["Income", "Income", "Expense"])
    cs.append(["Donations", "Fees", "General Expenses"])
    cs.append(["Individual Donation", "Monthly Member Fee", "Electricity"])
    cs.append(["Corporate Donation", "Annual Member Fee", "Office Supplies"])


# ---------------------------------------------------------------------------
# 1. valid-spanish.xlsx — Full valid workbook, Spanish headers
# ---------------------------------------------------------------------------


def gen_valid_spanish():
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    _add_metadata(ws)

    # Empty rows before header (mimics real Unicaja layout)
    for _ in range(4):
        ws.append([])

    ws.append(
        [
            "Fecha",
            "Valor",
            "Observaciones",
            "Importe",
            "Divisa",
            "Saldo",
            "Categoria",
            "Subcategoria",
            "Detalle",
        ]
    )

    ws.append(
        [
            "2025-01-02",
            "2025-01-02",
            "BULTO MILLET VICTOR",
            30,
            "EUR",
            71670.37,
            "Cuotas",
            "Cuota Socio Mensual",
            "transferencia",
        ]
    )
    ws.append(
        [
            "2025-01-03",
            "2025-01-03",
            "FRANCISCO BORJA",
            10,
            "EUR",
            71680.37,
            "Donaciones",
            "Donación Particular",
            "particular",
        ]
    )
    ws.append(
        [
            "2025-01-05",
            "2025-01-05",
            "ENDESA ENERGIA",
            -85.40,
            "EUR",
            71594.97,
            "Gastos Generales",
            "Electricidad",
            "recibo bimestral",
        ]
    )
    ws.append(
        [
            "2025-01-10",
            "2025-01-10",
            "EMPRESA SOLIDARIA SL",
            500,
            "EUR",
            72094.97,
            "Donaciones",
            "Donación Empresa",
            "RSC anual",
        ]
    )
    ws.append(
        [
            "2025-01-15",
            "2025-01-15",
            "AMAZON EU SARL",
            -42.99,
            "EUR",
            72051.98,
            "Gastos Generales",
            "Material Oficina",
            "pedido toner",
        ]
    )

    _add_categories_sheet_es(wb)
    _save(wb, "valid-spanish.xlsx")


# ---------------------------------------------------------------------------
# 2. valid-english.xlsx — Same data, English headers
# ---------------------------------------------------------------------------


def gen_valid_english():
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    _add_metadata(ws)

    for _ in range(4):
        ws.append([])

    ws.append(
        [
            "Date",
            "Value Date",
            "Description",
            "Amount",
            "Currency",
            "Balance",
            "Category",
            "Subcategory",
            "Detail",
        ]
    )

    ws.append(
        [
            "2025-01-02",
            "2025-01-02",
            "BULTO MILLET VICTOR",
            30,
            "EUR",
            71670.37,
            "Fees",
            "Monthly Member Fee",
            "transfer",
        ]
    )
    ws.append(
        [
            "2025-01-03",
            "2025-01-03",
            "FRANCISCO BORJA",
            10,
            "EUR",
            71680.37,
            "Donations",
            "Individual Donation",
            "individual",
        ]
    )
    ws.append(
        [
            "2025-01-05",
            "2025-01-05",
            "ENDESA ENERGIA",
            -85.40,
            "EUR",
            71594.97,
            "General Expenses",
            "Electricity",
            "bimonthly bill",
        ]
    )
    ws.append(
        [
            "2025-01-10",
            "2025-01-10",
            "EMPRESA SOLIDARIA SL",
            500,
            "EUR",
            72094.97,
            "Donations",
            "Corporate Donation",
            "annual CSR",
        ]
    )
    ws.append(
        [
            "2025-01-15",
            "2025-01-15",
            "AMAZON EU SARL",
            -42.99,
            "EUR",
            72051.98,
            "General Expenses",
            "Office Supplies",
            "toner order",
        ]
    )

    _add_categories_sheet_en(wb)
    _save(wb, "valid-english.xlsx")


# ---------------------------------------------------------------------------
# 3. with-duplicates.xlsx — Contains rows that'll match valid-spanish
# ---------------------------------------------------------------------------


def gen_with_duplicates():
    wb = Workbook()
    ws = wb.active
    ws.title = "UNICAJA 2026"
    _add_metadata(ws)

    for _ in range(4):
        ws.append([])

    ws.append(
        [
            "Fecha",
            "Valor",
            "Observaciones",
            "Importe",
            "Divisa",
            "Saldo",
            "Categoria",
            "Subcategoria",
            "Detalle",
        ]
    )

    # Duplicate of row 1 from valid-spanish
    ws.append(
        [
            "2025-01-02",
            "2025-01-02",
            "BULTO MILLET VICTOR",
            30,
            "EUR",
            71670.37,
            "Cuotas",
            "Cuota Socio Mensual",
            "transferencia",
        ]
    )
    # Duplicate of row 2 from valid-spanish
    ws.append(
        [
            "2025-01-03",
            "2025-01-03",
            "FRANCISCO BORJA",
            10,
            "EUR",
            71680.37,
            "Donaciones",
            "Donación Particular",
            "particular",
        ]
    )
    # Genuinely new transaction
    ws.append(
        [
            "2025-02-01",
            "2025-02-01",
            "NUEVA DONACION",
            25,
            "EUR",
            72076.98,
            "Donaciones",
            "Donación Particular",
            "nueva",
        ]
    )

    _add_categories_sheet_es(wb)
    _save(wb, "with-duplicates.xlsx")


# ---------------------------------------------------------------------------
# 4. data-errors.xlsx — Empty dates, bad amounts, missing categories
# ---------------------------------------------------------------------------


def gen_data_errors():
    wb = Workbook()
    ws = wb.active
    ws.title = "ERRORS TEST"
    _add_metadata(ws)

    ws.append(["Date", "Amount", "Category", "Subcategory", "Description"])

    # Valid row
    ws.append(["2025-01-01", 100, "Donations", "Individual Donation", "OK row"])
    # Empty date
    ws.append([None, 50, "Donations", "Individual Donation", "no date"])
    # Empty amount
    ws.append(["2025-01-03", None, "Donations", "Individual Donation", "no amount"])
    # Unparseable date
    ws.append(["not-a-date", 25, "Donations", "Individual Donation", "bad date"])
    # Empty category
    ws.append(["2025-01-05", 75, None, "Individual Donation", "no category"])
    # Empty subcategory
    ws.append(["2025-01-06", 60, "Donations", None, "no subcategory"])
    # Unparseable amount
    ws.append(["2025-01-07", "abc", "Donations", "Individual Donation", "bad amount"])

    _add_categories_sheet_en(wb)
    _save(wb, "data-errors.xlsx")


# ---------------------------------------------------------------------------
# 5. orphaned-subcategory.xlsx — Transaction references unknown subcategory
# ---------------------------------------------------------------------------


def gen_orphaned_subcategory():
    wb = Workbook()
    ws = wb.active
    ws.title = "ORPHAN TEST"
    _add_metadata(ws)

    ws.append(["Date", "Amount", "Category", "Subcategory"])

    # Valid row
    ws.append(["2025-01-01", 100, "Donations", "Individual Donation"])
    # Subcategory not in categories sheet
    ws.append(["2025-01-02", 50, "Donations", "Mystery Subcategory"])
    # Category not in categories sheet at all
    ws.append(["2025-01-03", 200, "Unknown Category", "Whatever"])

    cs = wb.create_sheet("Categories")
    cs.append(["Income"])
    cs.append(["Donations"])
    cs.append(["Individual Donation"])

    _save(wb, "orphaned-subcategory.xlsx")


# ---------------------------------------------------------------------------
# 6. missing-categories-sheet.xlsx — No categories sheet at all
# ---------------------------------------------------------------------------


def gen_missing_categories_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "NO CATS"
    _add_metadata(ws)

    ws.append(["Date", "Amount", "Category", "Subcategory"])
    ws.append(["2025-01-01", 100, "Donations", "Individual Donation"])

    # No second sheet created
    _save(wb, "missing-categories-sheet.xlsx")


# ---------------------------------------------------------------------------
# 7. missing-headers.xlsx — Bank mode (Date + Amount, no Category columns)
# ---------------------------------------------------------------------------


def gen_missing_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "BANK EXPORT"
    _add_metadata(ws)

    # Bank mode: Date + Amount (required) + Description, but NO Category/Subcategory
    ws.append(["Date", "Amount", "Description"])
    ws.append(["2025-01-01", 100, "Test transaction"])
    ws.append(["2025-01-02", -50, "Payment"])

    _save(wb, "missing-headers.xlsx")


# ---------------------------------------------------------------------------
# 8. empty-rows.xlsx — Valid headers but no data rows
# ---------------------------------------------------------------------------


def gen_empty_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "EMPTY DATA"
    _add_metadata(ws)

    ws.append(["Date", "Amount", "Category", "Subcategory"])
    # No data rows

    _add_categories_sheet_en(wb)
    _save(wb, "empty-rows.xlsx")


# ---------------------------------------------------------------------------
# 9. new-account.xlsx — IBAN that won't match any existing account
# ---------------------------------------------------------------------------


def gen_new_account():
    wb = Workbook()
    ws = wb.active
    ws.title = "BBVA 2026"
    _add_metadata(ws, iban="IBAN: ES91 0182 3456 7890 1234 5678", swift="BBVAESMMXXX")

    for _ in range(4):
        ws.append([])

    ws.append(["Date", "Amount", "Category", "Subcategory", "Description"])

    ws.append(["2025-03-01", 150, "Donations", "Individual Donation", "new bank transfer"])
    ws.append(["2025-03-05", -30, "General Expenses", "Office Supplies", "paper"])

    _add_categories_sheet_en(wb)
    _save(wb, "new-account.xlsx")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Generating import test fixtures...")
    gen_valid_spanish()
    gen_valid_english()
    gen_with_duplicates()
    gen_data_errors()
    gen_orphaned_subcategory()
    gen_missing_categories_sheet()
    gen_missing_headers()
    gen_empty_rows()
    gen_new_account()
    print(f"\nDone — {len(list(OUTPUT_DIR.glob('*.xlsx')))} fixtures in {OUTPUT_DIR}")
