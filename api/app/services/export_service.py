from __future__ import annotations

import io
from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font

if TYPE_CHECKING:
    from app.services.transaction_service import TransactionService


class ExportService:
    def __init__(self, *, transaction_service: TransactionService):
        self._transaction_service = transaction_service

    async def export_transactions_xlsx(
        self,
        date_from: date,
        date_to: date,
        account_id: str | None = None,
        category_id: str | None = None,
    ) -> bytes:
        items = await self._transaction_service.get_transactions_for_export(
            date_from=date_from.isoformat(),
            date_to=date_to.isoformat(),
            account_id=account_id,
            category_id=category_id,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        headers = [
            "Fecha",
            "Fecha Valor",
            "Cuenta",
            "Observaciones",
            "Categoría",
            "Subcategoría",
            "Etiquetas",
            "Importe",
            "Divisa",
            "Saldo",
            "Detalle",
            "Nº Mov",
            "Oficina",
        ]

        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font

        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.get("date"))
            ws.cell(row=row_idx, column=2, value=item.get("valueDate"))
            ws.cell(row=row_idx, column=3, value=item.get("accountId", ""))
            ws.cell(
                row=row_idx,
                column=4,
                value=item.get("bankDescription", ""),
            )
            ws.cell(row=row_idx, column=5, value=item.get("categoryId", ""))
            ws.cell(
                row=row_idx,
                column=6,
                value=item.get("subcategoryId", ""),
            )
            tag_ids = item.get("tagIds") or []
            ws.cell(row=row_idx, column=7, value=", ".join(tag_ids))
            amount_cell = ws.cell(
                row=row_idx,
                column=8,
                value=item.get("amount"),
            )
            amount_cell.number_format = "#,##0.00"
            ws.cell(
                row=row_idx,
                column=9,
                value=item.get("currency", "EUR"),
            )
            balance_cell = ws.cell(
                row=row_idx,
                column=10,
                value=item.get("balance"),
            )
            if balance_cell.value is not None:
                balance_cell.number_format = "#,##0.00"
            ws.cell(row=row_idx, column=11, value=item.get("detail", ""))
            ws.cell(
                row=row_idx,
                column=12,
                value=item.get("movementNumber", ""),
            )
            ws.cell(
                row=row_idx,
                column=13,
                value=item.get("branchNumber", ""),
            )

        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(2, min(len(items) + 2, 52)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
