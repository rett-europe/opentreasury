from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING
from unicodedata import normalize as unicode_normalize
from uuid import uuid4

from openpyxl import load_workbook

from app.constants.import_constants import (
    _ALIAS_LOOKUP,
    CATEGORY_HEADERS,
    CATEGORY_SHEET_NAMES,
    EXPENSE_ALIASES,
    INCOME_ALIASES,
    KNOWN_HEADERS,
    REQUIRED_HEADERS,
)
from app.models.domain import CategoryType, TransactionType
from app.models.schemas import CategoryCreate, SubcategoryCreate, TransactionCreate

if TYPE_CHECKING:
    from app.services.account_service import AccountService
    from app.services.category_service import CategoryService
    from app.services.transaction_service import TransactionService

logger = logging.getLogger(__name__)


@dataclass
class ImportSummary:
    import_batch_id: str
    import_mode: str
    account_id: str
    account_label: str
    categories_created: int = 0
    subcategories_added: int = 0
    transactions_imported: int = 0
    duplicates_skipped: int = 0
    rows_skipped: int = 0
    warnings: list[str] = field(default_factory=list)


class ImportService:
    def __init__(
        self,
        *,
        account_service: AccountService,
        category_service: CategoryService,
        transaction_service: TransactionService,
    ):
        self._accounts = account_service
        self._categories = category_service
        self._transactions = transaction_service

    # ------------------------------------------------------------------
    # Preview (dry-run validation)
    # ------------------------------------------------------------------

    async def preview_workbook(
        self,
        workbook_bytes: bytes,
        *,
        account_id: str,
        sheet: str | None = None,
        skip_duplicates: bool = True,
    ) -> dict:
        """Validate workbook without writing. Returns preview with valid flag, errors, and counts.

        When ``sheet`` is None and the workbook contains 2+ candidate movement sheets,
        a *sheet-selection-required* payload is returned so the caller can prompt the
        user to pick one. When ``sheet`` is provided, that specific sheet is validated.
        """
        errors: list[str] = []
        warnings: list[str] = []

        # 1. Validate account upfront (raises ValueError if missing/inactive)
        account_info = await self._validate_account(account_id)

        # 2. Open workbook
        try:
            workbook = self._load_workbook(workbook_bytes)
        except Exception:
            return self._preview_result(
                valid=False,
                errors=["Could not open file as a valid Excel workbook"],
                account=account_info,
            )

        # 3. Enumerate sheets — split into candidates and ignored
        candidates, ignored = self._enumerate_sheets(workbook)

        # 3a. If no sheet was requested and we have 2+ candidates → ask the user to pick one.
        if sheet is None and len(candidates) >= 2:
            return self._sheet_selection_result(
                candidates=candidates,
                ignored=ignored,
                account=account_info,
            )

        # 3b. Resolve the movement sheet:
        #     - If `sheet` is given, it must match an existing sheet (raise ValueError → 400 in router).
        #     - Otherwise pick the single candidate (or fall back to None for the zero-candidate error path).
        if sheet is not None:
            movement_sheet, header_row = self._select_sheet_by_name(workbook, sheet, candidates)
        elif candidates:
            movement_sheet = candidates[0]["sheet"]
            header_row = candidates[0]["header_row"]
        else:
            movement_sheet, header_row = None, None

        if movement_sheet is None:
            required = ", ".join(sorted(REQUIRED_HEADERS))
            error_msg = f"No sheet found with required column headers: {required}"
            return self._preview_result(
                valid=False,
                errors=[error_msg],
                account=account_info,
                ignored_sheets=ignored if not candidates else [],
                available_sheets=[c["name"] for c in candidates],
            )

        sheet_meta = {
            "selected_sheet": movement_sheet.title,
            "available_sheets": [c["name"] for c in candidates],
        }

        # 4. Build header map and validate required columns
        headers = self._build_header_map(movement_sheet, header_row)
        missing = REQUIRED_HEADERS - set(headers.keys())
        if missing:
            errors.append(f"Required columns not found: {', '.join(sorted(missing))}")
            return self._preview_result(valid=False, errors=errors, account=account_info, **sheet_meta)

        # 6. Detect import mode
        import_mode = self._detect_import_mode(workbook, headers)

        # 7. Validate data integrity — scan all data rows
        # Filter out trailing empty rows (openpyxl read_only mode can report
        # inflated max_row due to formatting-only cells).
        raw_rows = list(movement_sheet.iter_rows(min_row=header_row + 1, values_only=True))
        rows = [r for r in raw_rows if any(v not in (None, "") for v in r)]
        total_rows = len(rows)
        if total_rows == 0:
            return self._preview_result(
                valid=False,
                import_mode=import_mode,
                errors=["No data rows found below the header row"],
                account=account_info,
                **sheet_meta,
            )

        empty_dates: list[int] = []
        bad_dates: list[int] = []
        empty_amounts: list[int] = []
        bad_amounts: list[int] = []
        empty_categories: list[int] = []
        empty_subcategories: list[int] = []

        for idx, row in enumerate(rows, start=header_row + 1):
            d = self._cell(row, headers.get("date"))
            a = self._cell(row, headers.get("amount"))

            if not self._string_value(d):
                empty_dates.append(idx)
            elif self._to_date(d) is None:
                bad_dates.append(idx)

            if a is None or self._string_value(a) == "":
                empty_amounts.append(idx)
            elif self._to_decimal(a) is None:
                bad_amounts.append(idx)

            # Category validation only for Full and Inline modes
            if import_mode != "bank":
                c = self._cell(row, headers.get("category"))
                s = self._cell(row, headers.get("subcategory"))
                if not self._string_value(c):
                    empty_categories.append(idx)
                if not self._string_value(s):
                    empty_subcategories.append(idx)

        if empty_dates:
            errors.append(
                f"{len(empty_dates)} row(s) have an empty date — "
                f"check {self._format_rows(empty_dates)} in your Excel file"
            )
        if bad_dates:
            errors.append(
                f"{len(bad_dates)} row(s) have a date that could not be read — "
                f"check {self._format_rows(bad_dates)} in your Excel file. "
                f"Supported formats: DD/MM/YYYY, YYYY-MM-DD, or Excel date cells."
            )
        if empty_amounts:
            errors.append(
                f"{len(empty_amounts)} row(s) have an empty amount — "
                f"check {self._format_rows(empty_amounts)} in your Excel file"
            )
        if bad_amounts:
            errors.append(
                f"{len(bad_amounts)} row(s) have an amount that could not be read as a number — "
                f"check {self._format_rows(bad_amounts)} in your Excel file"
            )
        if import_mode != "bank":
            if empty_categories:
                errors.append(
                    f"{len(empty_categories)} row(s) have an empty category — "
                    f"check {self._format_rows(empty_categories)} in your Excel file"
                )
            if empty_subcategories:
                errors.append(
                    f"{len(empty_subcategories)} row(s) have an empty subcategory — "
                    f"check {self._format_rows(empty_subcategories)} in your Excel file"
                )

        rows_with_errors = len(empty_dates) + len(bad_dates) + len(empty_amounts) + len(bad_amounts)
        if import_mode != "bank":
            rows_with_errors += len(empty_categories) + len(empty_subcategories)

        if errors:
            return self._preview_result(
                valid=False,
                import_mode=import_mode,
                errors=errors,
                total_rows=total_rows,
                rows_with_errors=rows_with_errors,
                account=account_info,
                **sheet_meta,
            )

        # 8. Category / subcategory diff (mode-specific)
        new_categories: list[dict] = []
        new_subcategories: list[dict] = []
        resolvable: dict[str, set[str]] = {}

        existing_categories = await self._categories.list_categories()

        # Seed resolvable with existing DB categories
        for cat in existing_categories:
            cat_key = self._normalize_text(cat.get("name"))
            resolvable[cat_key] = {self._normalize_text(sub.get("name")) for sub in cat.get("subcategories", [])}

        if import_mode == "full":
            # Full mode: require categories sheet, diff against it
            category_sheet = self._find_categories_sheet(workbook)
            if category_sheet is None:
                names = ", ".join(sorted(CATEGORY_SHEET_NAMES))
                return self._preview_result(
                    valid=False,
                    import_mode=import_mode,
                    errors=[f"No categories sheet found. Expected a sheet named: {names}"],
                    account=account_info,
                    **sheet_meta,
                )
            parsed_categories = self._parse_category_sheet(category_sheet)
            new_categories, new_subcategories, resolvable, cat_warnings = self._compute_category_diff(
                parsed_categories, existing_categories, resolvable
            )
            warnings.extend(cat_warnings)

        elif import_mode == "inline":
            # Inline mode: extract categories from rows, resolve against DB
            extracted = self._extract_categories_from_rows(rows, headers)
            resolved, resolve_warnings = self._resolve_inline_categories(extracted, existing_categories)
            warnings.extend(resolve_warnings)
            new_categories, new_subcategories, resolvable, cat_warnings = self._compute_category_diff(
                resolved, existing_categories, resolvable
            )
            warnings.extend(cat_warnings)

        # Bank mode: no category diff — new_categories and new_subcategories stay empty

        # 9. Subcategory completeness — orphaned check (Full and Inline only)
        if import_mode != "bank":
            orphaned: set[str] = set()
            for row in rows:
                cat_text = self._string_value(self._cell(row, headers.get("category")))
                sub_text = self._string_value(self._cell(row, headers.get("subcategory")))
                if not cat_text or not sub_text:
                    continue
                cat_key = self._normalize_text(cat_text)
                sub_key = self._normalize_text(sub_text)
                if cat_key not in resolvable:
                    orphaned.add(f"Category '{cat_text}' not found in database or categories sheet")
                elif sub_key not in resolvable.get(cat_key, set()):
                    orphaned.add(f"Subcategory '{sub_text}' in '{cat_text}' not found in database or categories sheet")

            for msg in sorted(orphaned):
                errors.append(msg)

        if errors:
            return self._preview_result(
                valid=False,
                import_mode=import_mode,
                errors=errors,
                warnings=warnings,
                total_rows=total_rows,
                account=account_info,
                new_categories=new_categories,
                new_subcategories=new_subcategories,
                **sheet_meta,
            )

        # 10. Count transactions and duplicates
        valid_dates = [
            self._to_date(self._cell(row, headers.get("date")))
            for row in rows
            if self._to_date(self._cell(row, headers.get("date")))
        ]

        duplicates_to_skip = 0
        duplicate_rows: list[dict] = []
        transactions_to_import = 0

        if skip_duplicates and valid_dates and account_info.get("id"):
            existing_txns = await self._transactions.get_transactions_for_export(
                date_from=min(valid_dates).isoformat(),
                date_to=max(valid_dates).isoformat(),
                account_id=account_info["id"],
            )
            existing_keys = {
                self._transaction_identity(
                    item.get("date"),
                    item.get("bankDescription"),
                    item.get("detail"),
                    item.get("amount"),
                )
                for item in existing_txns
            }
            seen_keys = set(existing_keys)
            for row_offset, row in enumerate(rows, start=header_row + 1):
                tx_date = self._to_date(self._cell(row, headers.get("date")))
                amount = self._to_decimal(self._cell(row, headers.get("amount")))
                desc = self._string_value(self._cell(row, headers.get("description")))
                detail = self._build_detail(headers, row)
                identity = self._transaction_identity(
                    tx_date.isoformat() if tx_date else "",
                    desc,
                    detail,
                    amount,
                )
                if identity in seen_keys:
                    duplicates_to_skip += 1
                    duplicate_rows.append(
                        {
                            "row": row_offset,
                            "date": tx_date.isoformat() if tx_date else None,
                            "amount": float(amount) if amount is not None else None,
                            "description": self._truncate(desc, 120),
                        }
                    )
                else:
                    transactions_to_import += 1
                    seen_keys.add(identity)
        else:
            transactions_to_import = total_rows

        return self._preview_result(
            valid=True,
            import_mode=import_mode,
            errors=[],
            warnings=warnings,
            total_rows=total_rows,
            account=account_info,
            new_categories=new_categories,
            new_subcategories=new_subcategories,
            transactions_to_import=transactions_to_import,
            duplicates_to_skip=duplicates_to_skip,
            duplicate_rows=duplicate_rows,
            **sheet_meta,
        )

    def _preview_result(
        self,
        *,
        valid: bool,
        import_mode: str = "full",
        errors: list[str] | None = None,
        warnings: list[str] | None = None,
        total_rows: int = 0,
        rows_with_errors: int = 0,
        account: dict | None = None,
        new_categories: list[dict] | None = None,
        new_subcategories: list[dict] | None = None,
        transactions_to_import: int = 0,
        duplicates_to_skip: int = 0,
        duplicate_rows: list[dict] | None = None,
        selected_sheet: str | None = None,
        available_sheets: list[str] | None = None,
        ignored_sheets: list[dict] | None = None,
    ) -> dict:
        return {
            "valid": valid,
            "importMode": import_mode,
            "errors": errors or [],
            "warnings": warnings or [],
            "totalRows": total_rows,
            "rowsWithErrors": rows_with_errors,
            "account": account or {"id": "", "label": "", "iban": ""},
            "newCategories": new_categories or [],
            "newSubcategories": new_subcategories or [],
            "transactionsToImport": transactions_to_import,
            "duplicatesToSkip": duplicates_to_skip,
            "duplicateRows": duplicate_rows or [],
            "requiresSheetSelection": False,
            "selectedSheet": selected_sheet,
            "availableSheets": available_sheets or [],
            "ignoredSheets": ignored_sheets or [],
        }

    def _sheet_selection_result(
        self,
        *,
        candidates: list[dict],
        ignored: list[dict],
        account: dict,
    ) -> dict:
        """Return the discovery payload that asks the UI to prompt for a sheet."""
        return {
            "requiresSheetSelection": True,
            "candidateSheets": [
                {
                    "name": c["name"],
                    "dataRowCount": c["data_row_count"],
                    "headerRow": c["header_row"],
                }
                for c in candidates
            ],
            "ignoredSheets": ignored,
            "account": account,
            # Echo today's preview shape fields with safe defaults so callers that
            # branch on `requiresSheetSelection` still get well-formed JSON.
            "valid": False,
            "importMode": "full",
            "errors": [],
            "warnings": [],
            "totalRows": 0,
            "rowsWithErrors": 0,
            "newCategories": [],
            "newSubcategories": [],
            "transactionsToImport": 0,
            "duplicatesToSkip": 0,
            "duplicateRows": [],
            "selectedSheet": None,
            "availableSheets": [c["name"] for c in candidates],
        }

    async def _validate_account(self, account_id: str) -> dict:
        """Fetch and validate the selected account.

        Returns account dict with id, label, iban.
        Raises ValueError if not found or inactive.
        """
        account = await self._accounts.get_account(account_id)
        if not account or not account.get("isActive", True):
            raise ValueError(f"Account not found or inactive: {account_id}")
        return {
            "id": account["id"],
            "label": account.get("accountLabel", ""),
            "iban": account.get("iban", ""),
        }

    def _compute_category_diff(
        self,
        parsed_categories: list[dict],
        existing_categories: list[dict],
        resolvable: dict[str, set[str]],
    ) -> tuple[list[dict], list[dict], dict[str, set[str]], list[str]]:
        """Compute new categories/subcategories from parsed list against DB.

        Returns (new_categories, new_subcategories, updated_resolvable, warnings).
        """
        new_categories: list[dict] = []
        new_subcategories: list[dict] = []
        warnings: list[str] = []

        for sc in parsed_categories:
            cat_key = self._normalize_text(sc["name"])
            existing = self._find_category_by_name(existing_categories, sc["name"])

            if not existing:
                cat_type_val = sc["type"].value if sc["type"] else "expense"
                new_categories.append({"name": sc["name"], "type": cat_type_val, "suggestedType": "expense"})
                resolvable.setdefault(cat_key, set())
            else:
                if sc["type"] and existing.get("categoryType") != sc["type"].value:
                    warnings.append(
                        f"Category '{sc['name']}' exists with type '{existing.get('categoryType')}' "
                        f"but sheet declares it as '{sc['type'].value}'"
                    )

            for sub_name in sc["subcategories"]:
                sub_key = self._normalize_text(sub_name)
                if sub_key not in resolvable.get(cat_key, set()):
                    new_subcategories.append({"categoryName": sc["name"], "name": sub_name})
                    resolvable.setdefault(cat_key, set()).add(sub_key)

        return new_categories, new_subcategories, resolvable, warnings

    # ------------------------------------------------------------------
    # Import (commit)
    # ------------------------------------------------------------------

    async def import_workbook(
        self,
        workbook_bytes: bytes,
        *,
        account_id: str,
        category_type_overrides: dict[str, str] | None = None,
        user_id: str,
        user_name: str,
        sheet: str | None = None,
        skip_duplicates: bool = True,
    ) -> dict:
        workbook = self._load_workbook(workbook_bytes)

        candidates, _ = self._enumerate_sheets(workbook)
        if sheet is not None:
            movement_sheet, header_row = self._select_sheet_by_name(workbook, sheet, candidates)
        elif candidates:
            # Backwards compatibility: omit `sheet` → first-candidate-wins (today's behavior).
            movement_sheet = candidates[0]["sheet"]
            header_row = candidates[0]["header_row"]
        else:
            movement_sheet, header_row = None, None

        if movement_sheet is None:
            raise ValueError("No movement sheet found in workbook")

        headers = self._build_header_map(movement_sheet, header_row)
        import_mode = self._detect_import_mode(workbook, headers)
        import_batch_id = str(uuid4())
        import_source = f"excel-{import_mode}"

        account = await self._validate_account(account_id)
        summary = ImportSummary(
            import_batch_id=import_batch_id,
            import_mode=import_mode,
            account_id=account["id"],
            account_label=account["label"],
        )

        category_map: dict | None = None

        if import_mode == "full":
            category_sheet = self._find_categories_sheet(workbook)
            if category_sheet is None:
                raise ValueError("No categories sheet found in workbook")
            parsed = self._parse_category_sheet(category_sheet)
            category_map = await self._sync_categories(parsed, user_id=user_id, user_name=user_name, summary=summary)

        elif import_mode == "inline":
            raw_rows = list(movement_sheet.iter_rows(min_row=header_row + 1, values_only=True))
            rows = [r for r in raw_rows if any(v not in (None, "") for v in r)]
            existing_categories = await self._categories.list_categories()
            extracted = self._extract_categories_from_rows(rows, headers)
            resolved, resolve_warnings = self._resolve_inline_categories(extracted, existing_categories)
            summary.warnings.extend(resolve_warnings)
            category_map = await self._sync_categories(
                resolved,
                user_id=user_id,
                user_name=user_name,
                summary=summary,
                category_type_overrides=category_type_overrides,
            )

        # Bank mode: category_map stays None — no category sync

        await self._import_transactions(
            movement_sheet,
            header_row=header_row,
            headers=headers,
            account_id=account["id"],
            import_mode=import_mode,
            import_batch_id=import_batch_id,
            import_source=import_source,
            category_map=category_map,
            user_id=user_id,
            user_name=user_name,
            summary=summary,
            skip_duplicates=skip_duplicates,
        )

        return {
            "importBatchId": summary.import_batch_id,
            "importMode": summary.import_mode,
            "importSource": import_source,
            "accountId": summary.account_id,
            "accountLabel": summary.account_label,
            "selectedSheet": movement_sheet.title,
            "categoriesCreated": summary.categories_created,
            "subcategoriesAdded": summary.subcategories_added,
            "transactionsImported": summary.transactions_imported,
            "duplicatesSkipped": summary.duplicates_skipped,
            "rowsSkipped": summary.rows_skipped,
            "warnings": summary.warnings,
        }

    # ------------------------------------------------------------------
    # Mode detection
    # ------------------------------------------------------------------

    def _detect_import_mode(self, workbook, headers: dict[str, int]) -> str:
        """Detect import mode from workbook structure.

        Returns: "full", "inline", or "bank"
        """
        has_category_columns = CATEGORY_HEADERS.issubset(headers.keys())

        if not has_category_columns:
            return "bank"

        category_sheet = self._find_categories_sheet(workbook)
        if category_sheet is not None:
            return "full"

        return "inline"

    # ------------------------------------------------------------------
    # Category extraction (Inline mode)
    # ------------------------------------------------------------------

    def _extract_categories_from_rows(
        self,
        rows: list[tuple],
        headers: dict[str, int],
    ) -> list[dict]:
        """Extract unique (category, subcategory) pairs from transaction rows.

        Returns a list of dicts matching _parse_category_sheet() format:
        [{"name": str, "type": None, "subcategories": [str]}, ...]
        """
        categories: dict[str, dict] = {}  # normalized_name → {name, subcategories, subcategory_names}

        for row in rows:
            cat_text = self._string_value(self._cell(row, headers.get("category")))
            sub_text = self._string_value(self._cell(row, headers.get("subcategory")))

            if not cat_text or not sub_text:
                continue

            cat_key = self._normalize_text(cat_text)
            if cat_key not in categories:
                categories[cat_key] = {
                    "name": cat_text,
                    "subcategories": set(),
                    "subcategory_names": {},
                }

            sub_key = self._normalize_text(sub_text)
            if sub_key not in categories[cat_key]["subcategory_names"]:
                categories[cat_key]["subcategories"].add(sub_key)
                categories[cat_key]["subcategory_names"][sub_key] = sub_text

        return [
            {
                "name": info["name"],
                "type": None,
                "subcategories": list(info["subcategory_names"].values()),
            }
            for info in categories.values()
        ]

    def _resolve_inline_categories(
        self,
        extracted: list[dict],
        existing_categories: list[dict],
    ) -> tuple[list[dict], list[str]]:
        """Resolve extracted categories against DB. Assign types.

        Returns: (resolved_list, warnings)
        """
        resolved: list[dict] = []
        warnings: list[str] = []

        for item in extracted:
            existing = self._find_category_by_name(existing_categories, item["name"])
            if existing:
                cat_type = CategoryType(existing.get("categoryType", "expense"))
            else:
                cat_type = CategoryType.EXPENSE
            resolved.append(
                {
                    "name": item["name"],
                    "type": cat_type,
                    "subcategories": item["subcategories"],
                }
            )

        return resolved, warnings

    # ------------------------------------------------------------------
    # Transaction type inference
    # ------------------------------------------------------------------

    @staticmethod
    def _infer_transaction_type(amount: Decimal) -> TransactionType:
        """Infer transaction type from amount sign.

        Positive/zero → income, negative → expense.
        """
        if amount >= 0:
            return TransactionType.INCOME
        return TransactionType.EXPENSE

    # ------------------------------------------------------------------
    # Workbook loading
    # ------------------------------------------------------------------

    @staticmethod
    def _load_workbook(workbook_bytes: bytes):
        """Open workbook with openpyxl, falling back to read-only mode.

        Some bank-exported workbooks contain pivot tables that trigger a bug
        in openpyxl <= 3.1.x (``Nested.from_tree()`` missing argument).  When
        the standard load fails we retry with ``read_only=True`` which skips
        pivot-cache parsing.  A warning is logged so we notice if this path is
        hit frequently.
        """
        try:
            return load_workbook(BytesIO(workbook_bytes), data_only=True)
        except Exception:
            logger.warning("Standard workbook load failed; retrying with read_only=True")
            return load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=True)

    # ------------------------------------------------------------------
    # Sheet / header detection (multilingual)
    # ------------------------------------------------------------------

    def _find_movements_sheet(self, workbook):
        """Find the first sheet with required headers. Returns (sheet, header_row) or (None, None).

        Kept for backward compatibility with any external callers; new code paths use
        ``_enumerate_sheets`` + ``_select_sheet_by_name`` so they can offer the user a choice.
        """
        for sheet in workbook.worksheets:
            header_row = self._find_header_row(sheet)
            if header_row is not None:
                return sheet, header_row
        return None, None

    def _enumerate_sheets(self, workbook) -> tuple[list[dict], list[dict]]:
        """Return ``(candidates, ignored)`` for every worksheet in the workbook.

        Each candidate dict contains: ``name``, ``sheet``, ``header_row``, ``data_row_count``.
        Each ignored dict contains: ``name``, ``reason``, and (when applicable) ``missing``.
        Order is preserved from workbook order — the caller can rely on it for default selection.
        """
        candidates: list[dict] = []
        ignored: list[dict] = []
        for sheet in workbook.worksheets:
            header_row = self._find_header_row(sheet)
            if header_row is None:
                # Distinguish "empty sheet" from "has rows but missing headers" so we can
                # show a useful reason in the UI without needing to compute the full diff.
                if sheet.max_row is None or sheet.max_row <= 0 or self._sheet_is_empty(sheet):
                    ignored.append({"name": sheet.title, "reason": "empty"})
                else:
                    # Compute which required headers were not seen anywhere in the
                    # candidate header zone (first 12 rows), so the UI can tell the
                    # user *which* columns to add — not just that some are missing.
                    # (`missing` cannot be empty here: if both required headers were
                    # present in any of the first 12 rows, `_find_header_row` would
                    # have succeeded and we wouldn't be in this branch.)
                    found = self._collect_known_headers(sheet)
                    missing = sorted(REQUIRED_HEADERS - found)
                    ignored.append(
                        {
                            "name": sheet.title,
                            "reason": "missing_required_headers",
                            "missing": missing,
                        }
                    )
                continue

            # Cheap data-row count: max_row minus header_row (clamped at zero).
            # openpyxl's max_row is already cached, so this is O(1).
            data_row_count = max(0, (sheet.max_row or 0) - header_row)
            candidates.append(
                {
                    "name": sheet.title,
                    "sheet": sheet,
                    "header_row": header_row,
                    "data_row_count": data_row_count,
                }
            )
        return candidates, ignored

    def _select_sheet_by_name(self, workbook, name: str, candidates: list[dict]):
        """Resolve a user-selected sheet name to ``(sheet, header_row)``.

        Raises ValueError (mapped to HTTP 400 by the router) when:
          - the sheet name does not exist in the workbook, or
          - the sheet exists but is not a valid movements candidate.
        """
        if name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' not found in workbook")
        for candidate in candidates:
            if candidate["name"] == name:
                return candidate["sheet"], candidate["header_row"]
        raise ValueError(f"Sheet '{name}' is not importable: missing required headers")

    @staticmethod
    def _sheet_is_empty(sheet) -> bool:
        """Return True if the sheet has no non-empty cell value.

        The check is intentionally bounded so sheet discovery stays cheap even
        for worksheets with very large used ranges. We scan only an initial
        window of cells; if the worksheet fits entirely inside that window and
        no values are found, it is empty. For larger worksheets, absence of a
        value in the sampled region is treated conservatively as non-empty.
        """
        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        if max_row == 0 or max_col == 0:
            return True

        row_limit = min(max_row, 50)
        col_limit = min(max_col, 20)

        for row in sheet.iter_rows(
            min_row=1,
            max_row=row_limit,
            min_col=1,
            max_col=col_limit,
            values_only=True,
        ):
            for value in row:
                if value not in (None, ""):
                    return False

        return max_row <= row_limit and max_col <= col_limit

    def _collect_known_headers(self, sheet) -> set[str]:
        """Return the set of canonical header keys found in the first 12 rows.

        Used by ``_enumerate_sheets`` to report which *specific* required
        headers are missing on a non-candidate sheet.
        Uses ``iter_rows`` instead of ``sheet[row_idx]`` for read-only compatibility.
        """
        found: set[str] = set()
        scan_limit = min(sheet.max_row or 0, 12)
        if scan_limit <= 0:
            return found
        for row in sheet.iter_rows(min_row=1, max_row=scan_limit):
            for cell in row:
                if cell.value is None:
                    continue
                canonical = self._resolve_alias(cell.value)
                if canonical:
                    found.add(canonical)
        return found

    def _find_categories_sheet(self, workbook):
        """Find the categories sheet by name. Returns sheet or None."""
        for sheet in workbook.worksheets:
            if self._normalize_text(sheet.title) in CATEGORY_SHEET_NAMES:
                return sheet
        return None

    def _find_header_row(self, sheet) -> int | None:
        """Scan the first 12 rows for one that contains all required canonical headers.

        Uses ``iter_rows`` instead of ``sheet[row_idx]`` for read-only compatibility.
        """
        scan_limit = min(sheet.max_row or 0, 12)
        if scan_limit <= 0:
            return None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=scan_limit), start=1):
            found_keys: set[str] = set()
            for cell in row:
                if cell.value is None:
                    continue
                canonical = self._resolve_alias(cell.value)
                if canonical in REQUIRED_HEADERS:
                    found_keys.add(canonical)
            if REQUIRED_HEADERS.issubset(found_keys):
                return row_idx
        return None

    def _build_header_map(self, sheet, header_row: int) -> dict[str, int]:
        """Map canonical column keys to column indices.

        Known columns are mapped to their canonical key (e.g., 'fecha' → 'date').
        Unrecognized columns are kept under their normalized header text so they
        can be captured into the detail field by _build_detail.
        """
        result: dict[str, int] = {}
        for row in sheet.iter_rows(min_row=header_row, max_row=header_row):
            for idx, cell in enumerate(row):
                if cell.value is None:
                    continue
                canonical = self._resolve_alias(cell.value)
                if canonical and canonical not in result:
                    result[canonical] = idx
                elif not canonical:
                    # Unrecognized column — normalize and keep for detail capture
                    raw = self._normalize_header(cell.value)
                    if raw and raw not in result:
                        result[raw] = idx
        return result

    def _resolve_alias(self, value) -> str | None:
        """Normalize a header cell and resolve to a canonical key."""
        normalized = self._normalize_header(value)
        if not normalized:
            return None
        canonical = _ALIAS_LOOKUP.get(normalized)
        if canonical:
            return canonical
        return None

    def _normalize_header(self, value) -> str:
        """Normalize a header cell for alias matching."""
        text = self._normalize_text(value)
        if text.startswith("numero "):
            text = "n " + text[7:]
        return text

    def _parse_category_sheet(self, sheet) -> list[dict]:
        """Extract [{name, type: CategoryType, subcategories: [str]}] from categories sheet."""
        result = []
        for col in range(1, sheet.max_column + 1):
            cat_type = self._category_type_from_sheet(sheet.cell(row=1, column=col).value)
            cat_name = self._string_value(sheet.cell(row=2, column=col).value)
            if not cat_type or not cat_name:
                continue
            subcategories = []
            for row in range(3, sheet.max_row + 1):
                sub_name = self._string_value(sheet.cell(row=row, column=col).value)
                if sub_name:
                    subcategories.append(sub_name)
            result.append({"name": cat_name, "type": cat_type, "subcategories": subcategories})
        return result

    def _category_type_from_sheet(self, value) -> CategoryType | None:
        normalized = self._normalize_text(value)
        if normalized in INCOME_ALIASES:
            return CategoryType.INCOME
        if normalized in EXPENSE_ALIASES:
            return CategoryType.EXPENSE
        return None

    # ------------------------------------------------------------------
    # Category sync (creates missing categories / subcategories)
    # ------------------------------------------------------------------

    async def _sync_categories(
        self,
        parsed_categories: list[dict],
        *,
        user_id: str,
        user_name: str,
        summary: ImportSummary,
        category_type_overrides: dict[str, str] | None = None,
    ) -> dict:
        existing_categories = await self._categories.list_categories()
        category_map: dict = {}

        for cat_info in parsed_categories:
            category = self._find_category_by_name(existing_categories, cat_info["name"])
            if not category:
                resolved_type = cat_info["type"]
                if category_type_overrides and cat_info["name"] in category_type_overrides:
                    resolved_type = CategoryType(category_type_overrides[cat_info["name"]])
                category = await self._categories.create_category(
                    CategoryCreate(
                        name=cat_info["name"],
                        category_type=resolved_type,
                        sort_order=len(existing_categories),
                        subcategories=[],
                    ),
                    user_id=user_id,
                    user_name=user_name,
                )
                existing_categories.append(category)
                summary.categories_created += 1
            elif cat_info["type"] and category.get("categoryType") != cat_info["type"].value:
                summary.warnings.append(
                    f"Category '{cat_info['name']}' exists with type '{category.get('categoryType')}' "
                    f"but sheet declares '{cat_info['type'].value}'"
                )

            sub_map = {
                self._normalize_text(sub.get("name")): sub.get("id") for sub in category.get("subcategories", [])
            }

            for sub_name in cat_info["subcategories"]:
                key = self._normalize_text(sub_name)
                if key in sub_map:
                    continue
                category = await self._categories.add_subcategory(
                    category_id=category["id"],
                    data=SubcategoryCreate(name=sub_name),
                    user_id=user_id,
                    user_name=user_name,
                )
                summary.subcategories_added += 1
                sub_map = {
                    self._normalize_text(sub.get("name")): sub.get("id") for sub in category.get("subcategories", [])
                }

            category_map[self._normalize_text(cat_info["name"])] = {
                "id": category["id"],
                "type": category.get("categoryType"),
                "subcategories": {
                    self._normalize_text(sub.get("name")): sub.get("id") for sub in category.get("subcategories", [])
                },
            }

        return category_map

    # ------------------------------------------------------------------
    # Transaction import
    # ------------------------------------------------------------------

    async def _import_transactions(
        self,
        sheet,
        *,
        header_row: int,
        headers: dict[str, int],
        account_id: str,
        import_mode: str,
        import_batch_id: str,
        import_source: str,
        category_map: dict | None,
        user_id: str,
        user_name: str,
        summary: ImportSummary,
        skip_duplicates: bool = True,
    ) -> None:
        raw_rows = list(sheet.iter_rows(min_row=header_row + 1, values_only=True))
        rows = [r for r in raw_rows if any(v not in (None, "") for v in r)]
        if not rows:
            return

        dates = [
            self._to_date(self._cell(row, headers.get("date")))
            for row in rows
            if self._to_date(self._cell(row, headers.get("date")))
        ]
        if not dates:
            return

        existing = await self._transactions.get_transactions_for_export(
            date_from=min(dates).isoformat(),
            date_to=max(dates).isoformat(),
            account_id=account_id,
        )
        existing_keys = {
            self._transaction_identity(
                item.get("date"),
                item.get("bankDescription"),
                item.get("detail"),
                item.get("amount"),
            )
            for item in existing
        }
        imported_keys = set(existing_keys)

        for offset, row in enumerate(rows, start=header_row + 1):
            tx_date = self._to_date(self._cell(row, headers.get("date")))
            amount = self._to_decimal(self._cell(row, headers.get("amount")))

            if not tx_date or amount is None:
                summary.rows_skipped += 1
                summary.warnings.append(f"Row {offset}: missing date, category, subcategory, or amount.")
                continue

            # Category resolution (mode-dependent)
            category_id: str | None = None
            subcategory_id: str | None = None

            if import_mode != "bank":
                category_name = self._string_value(self._cell(row, headers.get("category")))
                subcategory_name = self._string_value(self._cell(row, headers.get("subcategory")))

                if not category_name or not subcategory_name:
                    summary.rows_skipped += 1
                    summary.warnings.append(f"Row {offset}: missing date, category, subcategory, or amount.")
                    continue

                if category_map is not None:
                    category_info = category_map.get(self._normalize_text(category_name))
                    if not category_info:
                        summary.rows_skipped += 1
                        summary.warnings.append(f"Row {offset}: category '{category_name}' not found.")
                        continue

                    subcategory_id = category_info["subcategories"].get(self._normalize_text(subcategory_name))
                    if not subcategory_id:
                        summary.rows_skipped += 1
                        summary.warnings.append(
                            f"Row {offset}: subcategory '{subcategory_name}' not found in '{category_name}'."
                        )
                        continue
                    category_id = category_info["id"]

            # Duplicate detection
            bank_description = self._truncate(
                self._string_value(row[headers["description"]]) if "description" in headers else None, 500
            )
            detail = self._build_detail(headers, row)
            identity = self._transaction_identity(
                tx_date.isoformat(),
                bank_description,
                detail,
                amount,
            )
            if identity in imported_keys and skip_duplicates:
                summary.duplicates_skipped += 1
                continue

            # Build transaction fields
            value_date = self._to_date(row[headers["value_date"]]) if "value_date" in headers else tx_date
            currency = self._string_value(row[headers["currency"]]) if "currency" in headers else "EUR"
            balance = self._to_decimal(row[headers["balance"]]) if "balance" in headers else None

            # Transaction type from amount sign (all modes)
            tx_type = self._infer_transaction_type(amount)

            await self._transactions.create_transaction(
                TransactionCreate(
                    transaction_date=tx_date,
                    value_date=value_date,
                    amount=amount,
                    currency=currency or "EUR",
                    bank_description=bank_description,
                    account_id=account_id,
                    transaction_type=tx_type,
                    category_id=category_id,
                    subcategory_id=subcategory_id,
                    detail=detail,
                    balance=balance,
                    tag_ids=[],
                    import_batch_id=import_batch_id,
                    import_source=import_source,
                ),
                user_id=user_id,
                user_name=user_name,
            )
            imported_keys.add(identity)
            summary.transactions_imported += 1

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _format_rows(row_numbers: list[int], max_display: int = 10) -> str:
        """Format row numbers for error messages. Shows first N rows, then '...'."""
        if len(row_numbers) == 1:
            return f"row {row_numbers[0]}"
        if len(row_numbers) <= max_display:
            return "rows " + ", ".join(str(r) for r in row_numbers)
        shown = ", ".join(str(r) for r in row_numbers[:max_display])
        return f"rows {shown} and {len(row_numbers) - max_display} more"

    def _build_detail(self, headers: dict[str, int], row) -> str | None:
        parts: list[str] = []
        detail_idx = headers.get("detail")
        if detail_idx is not None:
            val = self._string_value(row[detail_idx])
            if val:
                parts.append(val)

        # Capture any unrecognized columns (not in KNOWN_HEADERS) into detail
        for key, idx in headers.items():
            if key in KNOWN_HEADERS:
                continue
            val = self._string_value(row[idx])
            if val:
                parts.append(f"{key}: {val}")

        result = " | ".join(parts) if parts else None
        return self._truncate(result, 2000)

    @staticmethod
    def _truncate(value: str | None, max_len: int) -> str | None:
        """Truncate a string to max_len, appending '...' if truncated."""
        if value is None or len(value) <= max_len:
            return value
        return value[: max_len - 3] + "..."

    def _find_category_by_name(self, categories: list[dict], name: str) -> dict | None:
        normalized = self._normalize_text(name)
        for category in categories:
            if self._normalize_text(category.get("name")) == normalized:
                return category
        return None

    def _transaction_identity(self, tx_date, bank_description, detail, amount) -> str:
        amount_decimal = self._to_decimal(amount)
        amount_value = str(abs(amount_decimal)) if amount_decimal is not None else ""
        return "|".join(
            [
                self._string_value(tx_date),
                self._normalize_text(bank_description),
                self._normalize_text(detail),
                amount_value,
            ]
        )

    def _to_date(self, value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            raw = value.strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(raw).date()
            except ValueError:
                return None
        return None

    def _to_decimal(self, value) -> Decimal | None:
        if value is None or value == "":
            return None
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except Exception:
            return None

    def _string_value(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_text(self, value) -> str:
        text = self._string_value(value).lower()
        text = unicode_normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
        text = text.replace("n\u00b0", "n").replace("/", " ").replace("-", " ")
        return " ".join(text.split())

    @staticmethod
    def _cell(row: tuple, index: int | None):
        """Safely get a cell value from a row tuple."""
        if index is None or index < 0 or index >= len(row):
            return None
        return row[index]
